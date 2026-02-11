#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
ECH (China) SunSirs -> Excel

Correção Crítica:
- Ajuste de formatação de moeda (USD e RMB).
- O site usa formato internacional (12,500.00), onde vírgula é milhar.
- Script ajustado para não multiplicar por 100 erroneamente.

Deps:
  pip install --upgrade pandas requests beautifulsoup4 openpyxl yfinance playwright
  playwright install chromium
"""

import os
import re
import sys
import time
import argparse
from datetime import datetime, timedelta, date
from pathlib import Path

import pandas as pd
import requests
import yfinance as yf
from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


# ========================= CONFIG =========================

SUNSIRS_URLS = [
    "https://www.sunsirs.com/uk/prodetail-439.html",
    "https://www.sunsirs.com/prodetail-439.html",
]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0 Safari/537.36"
)

SHEET_NAME = "ECH_Daily"


# ========================= OneDrive dinâmico =========================

def _one_drive_base() -> Path:
    for var in ("OneDriveCommercial", "OneDriveConsumer", "OneDrive"):
        p = os.environ.get(var)
        if p and Path(p).exists():
            return Path(p)
    home = Path.home()
    if (home / "OneDrive").exists():
        return home / "OneDrive"
    return home / "Documents"


BASE_DIR = _one_drive_base() / "Arquivos" / "Python" / "excel-api_restful"
BASE_DIR.mkdir(parents=True, exist_ok=True)

OUT_XLSX = BASE_DIR / "pogo_spread_history.xlsx"
DEBUG_HTML = BASE_DIR / "debug_sunsirs_ech.html"


# ========================= Utils =========================

def log(msg: str, quiet: bool = False):
    if not quiet:
        print(msg, flush=True)


def to_ts(col) -> pd.Series:
    arr = pd.to_datetime(col, errors="coerce", utc=False)
    if isinstance(arr, (pd.DatetimeIndex, pd.Index)) or not hasattr(arr, "dt"):
        arr = pd.Series(arr)
    try:
        arr = arr.dt.tz_convert("UTC").dt.tz_localize(None)
    except Exception:
        pass
    return arr.dt.normalize()


def to_date_str(col) -> pd.Series:
    arr = pd.to_datetime(col, errors="coerce", utc=False)
    if isinstance(arr, (pd.DatetimeIndex, pd.Index)) or not hasattr(arr, "dt"):
        arr = pd.Series(arr)
    try:
        arr = arr.dt.tz_convert("UTC").dt.tz_localize(None)
    except Exception:
        pass
    return arr.dt.normalize().dt.strftime("%Y-%m-%d")


# ========================= Excel (preserva abas) =========================

def write_df_to_excel(path: Path, sheet_name: str, df: pd.DataFrame):
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if wb.worksheets:
            wb.remove(wb.worksheets[0])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wb.save(path)


def load_existing() -> pd.DataFrame:
    if OUT_XLSX.exists():
        try:
            df = pd.read_excel(
                OUT_XLSX, sheet_name=SHEET_NAME, engine="openpyxl")
            if "Date" in df:
                df["Date"] = to_ts(df["Date"])
            if "FX_used_date" in df:
                df["FX_used_date"] = to_date_str(df["FX_used_date"])
            if "Date" in df:
                df = df.drop_duplicates(subset=["Date"])
            return df
        except Exception:
            pass

    return pd.DataFrame(columns=["Date", "Price_RMB_ton", "Price_USD_ton", "FX_source", "FX_used_date"])


# ========================= HTML Fetch =========================

def fetch_html_requests(url: str) -> str | None:
    try:
        r = requests.get(
            url,
            headers={
                "User-Agent": USER_AGENT,
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
            timeout=(10, 40),
        )
        r.raise_for_status()
        return r.text
    except Exception:
        return None


def fetch_html_playwright(url: str, quiet: bool = False) -> str:
    log(f"[Playwright] Abrindo: {url}", quiet)

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled",
                  "--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = browser.new_context(
            user_agent=USER_AGENT,
            locale="en-US",
            viewport={"width": 1366, "height": 768},
        )
        page = context.new_page()
        page.set_default_timeout(45000)

        try:
            page.goto(url, wait_until="domcontentloaded")
            try:
                page.wait_for_load_state("networkidle", timeout=45000)
            except PWTimeoutError:
                pass

            page.wait_for_timeout(1500)
            html = page.content()
        finally:
            context.close()
            browser.close()

    return html


# ========================= Parser CORRIGIDO =========================

def extract_ech_from_html(html: str) -> pd.DataFrame:
    if not html or len(html) < 200:
        return pd.DataFrame(columns=["Date", "Price_RMB_ton"])

    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(" ", strip=True)

    # Regex para capturar preço e data
    pat = re.compile(
        r"ECH.*?(?:Chemical|Produtos químicos).*?([\d.,]+)\s+(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})",
        re.IGNORECASE
    )

    rows = []
    for m in pat.finditer(text):
        price_str = m.group(1)
        d_str = m.group(2)

        try:
            # === CORREÇÃO DE FORMATAÇÃO ===
            # Formato esperado: 12,600.00 (Inglês/Internacional)
            # Ação: Remover vírgula, manter ponto.
            
            clean_price = price_str.replace(",", "") # Remove separador de milhar
            # Se sobrar algum ponto extra no final (ex: "12600."), o float resolve.
            
            price = float(clean_price)

            # === TRAVA DE SEGURANÇA ===
            # O preço do ECH gira em torno de 8.000 a 20.000 RMB.
            # Se o preço vier > 100.000, significa que o ponto decimal foi ignorado.
            if price > 100000:
                price = price / 100.0

            # Tratamento de Data
            if "/" in d_str:
                dt = datetime.strptime(d_str, "%d/%m/%Y")
            else:
                dt = datetime.strptime(d_str, "%Y-%m-%d")

            # Filtro de valor mínimo (evita zeros ou lixo)
            if price > 100:
                rows.append({"Date": dt, "Price_RMB_ton": price})
        
        except Exception as e:
            continue

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Date", "Price_RMB_ton"])

    df["Date"] = to_ts(df["Date"])
    df = df.drop_duplicates(subset=["Date"]).sort_values(
        "Date", ascending=False).reset_index(drop=True)
    return df


def fetch_ech_rows(quiet: bool = False) -> pd.DataFrame:
    last_html = None

    # 1) Requests
    for url in SUNSIRS_URLS:
        html = fetch_html_requests(url)
        if html:
            last_html = html
            df = extract_ech_from_html(html)
            if not df.empty:
                log(f"[SunSirs] OK via requests | {url} | linhas: {len(df)}", quiet)
                return df

    # 2) Playwright fallback
    for url in SUNSIRS_URLS:
        html = fetch_html_playwright(url, quiet=quiet)
        last_html = html
        df = extract_ech_from_html(html)
        if not df.empty:
            log(f"[SunSirs] OK via Playwright | {url} | linhas: {len(df)}", quiet)
            return df

    # Debug
    if last_html:
        try:
            DEBUG_HTML.write_text(last_html, encoding="utf-8")
        except Exception:
            pass

    raise RuntimeError(
        "Não foi possível extrair histórico de ECH da SunSirs.")


# ========================= FX RMB -> USD =========================

def _fx_exchangerate_host(dt_: date) -> tuple[float | None, date | None]:
    # Ex: 1 USD = 7.20 CNY -> Rate deve ser ~0.138
    url = f"https://api.exchangerate.host/{dt_.isoformat()}?base=CNY&symbols=USD"
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=(8, 25))
    r.raise_for_status()
    data = r.json()
    rate = data.get("rates", {}).get("USD")
    return (float(rate), dt_) if rate else (None, None)


def _fx_yahoo(dt_: date, lookback_days: int) -> tuple[float | None, date | None]:
    start = pd.Timestamp(dt_) - pd.Timedelta(days=lookback_days + 3)
    end = pd.Timestamp(dt_) + pd.Timedelta(days=1)
    
    # USDCNY=X retorna quantos RMB valem 1 USD (ex: 7.25)
    hist = yf.Ticker("USDCNY=X").history(
        start=start, end=end, auto_adjust=False, actions=False)

    if hist is None or hist.empty:
        return (None, None)

    idx = hist.index
    if getattr(idx, "tz", None) is not None:
        idx = idx.tz_convert("UTC").tz_localize(None)

    idx_norm = pd.DatetimeIndex(idx).normalize()
    cutoff = pd.Timestamp(dt_)
    mask = idx_norm <= cutoff
    if not mask.any():
        return (None, None)

    last_pos = mask.nonzero()[0][-1]
    close_val = float(hist["Close"].iloc[last_pos])
    
    if close_val <= 0:
        return (None, None)

    # Inverte para ter o fator multiplicador (CNY -> USD)
    # Se 1 USD = 7.25 CNY, então 1 CNY = 1/7.25 = 0.1379 USD
    cny_to_usd = 1.0 / close_val
    used_day = idx_norm[last_pos].date()
    return (cny_to_usd, used_day)


def fetch_cny_usd_rate_for_date(dt_any, max_lookback: int = 30) -> tuple[float, str, date]:
    if isinstance(dt_any, pd.Timestamp):
        base_date = dt_any.date()
    elif isinstance(dt_any, datetime):
        base_date = dt_any.date()
    else:
        base_date = dt_any

    for i in range(max_lookback + 1):
        day = base_date - timedelta(days=i)
        try:
            rate, used_day = _fx_exchangerate_host(day)
            if rate:
                return (rate, "exchangerate.host", used_day)
        except Exception:
            continue

    rate, used_day = _fx_yahoo(base_date, lookback_days=max_lookback)
    if rate:
        return (rate, "yahoo", used_day)

    raise RuntimeError(
        f"Sem taxa CNY->USD para {base_date} (lookback={max_lookback}).")


def add_usd_column(df: pd.DataFrame, lookback: int, quiet: bool = False) -> pd.DataFrame:
    usd_prices, sources, used_days = [], [], []
    n = len(df)

    for i, (dt_ts, price_rmb) in enumerate(zip(df["Date"], df["Price_RMB_ton"]), start=1):
        time.sleep(0.10)
        rate, src, used_day = fetch_cny_usd_rate_for_date(
            dt_ts, max_lookback=lookback)

        # Cálculo final: Preço corrigido (RMB) * Taxa
        usd_val = float(price_rmb) * float(rate)
        
        usd_prices.append(usd_val)
        sources.append(src)
        used_days.append(used_day)

        log(f"[{i}/{n}] {dt_ts.date()} | RMB: {price_rmb} -> USD: {usd_val:.2f}", quiet)

    out = df.copy()
    out["Price_USD_ton"] = usd_prices
    out["FX_source"] = sources
    out["FX_used_date"] = to_date_str(used_days)
    out["Date"] = to_ts(out["Date"])
    return out


# ========================= Main =========================

def parse_args():
    p = argparse.ArgumentParser(
        description="ECH China -> Excel (SunSirs + FX RMB->USD)")
    p.add_argument("--limit", type=int, default=30,
                   help="quantas datas recentes buscar (default: 30)")
    p.add_argument("--lookback", type=int, default=30,
                   help="lookback do FX (default: 30)")
    p.add_argument("--quiet", action="store_true", help="suprime logs")
    return p.parse_args()


def main():
    args = parse_args()

    try:
        # Busca dados e converte datas
        df_recent = fetch_ech_rows(quiet=args.quiet).head(max(1, args.limit))
        df_recent["Date"] = to_ts(df_recent["Date"])

        # Carrega existente
        df_old = load_existing()
        existing_dates = set()
        if not df_old.empty and "Date" in df_old:
            existing_dates = set(df_old["Date"].dropna().dt.normalize())

        # Filtra apenas novos
        df_new = df_recent[~df_recent["Date"].isin(existing_dates)].copy()

        if df_new.empty:
            log("Nenhuma data nova encontrada.", args.quiet)
            print(f"OK: {OUT_XLSX} já atualizada.")
            return 0

        log(f"Novas linhas: {len(df_new)}. Calculando conversão USD...", args.quiet)

        # Adiciona coluna USD com conversão correta
        df_new = add_usd_column(df_new, lookback=max(1, args.lookback), quiet=args.quiet)

        # Concatena
        frames = []
        if df_old is not None and not df_old.empty:
            frames.append(df_old.dropna(how="all"))
        if df_new is not None and not df_new.empty:
            frames.append(df_new.dropna(how="all"))

        df_all = pd.concat(frames, ignore_index=True) if frames else df_new

        # Ordenação e limpeza final
        df_all["Date"] = to_ts(df_all["Date"])
        df_all["FX_used_date"] = to_date_str(df_all["FX_used_date"]) if "FX_used_date" in df_all else None
        df_all = df_all.drop_duplicates(subset=["Date"]).sort_values("Date").reset_index(drop=True)

        write_df_to_excel(OUT_XLSX, SHEET_NAME, df_all)

        print(f"SUCESSO: Arquivo atualizado -> {OUT_XLSX}")
        return 0

    except Exception as e:
        print(f"ERRO FATAL: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())